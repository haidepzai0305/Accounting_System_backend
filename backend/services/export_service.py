import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def build_excel(title: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    """
    Build a styled Excel workbook and return it as a BytesIO stream.

    :param title:   Sheet name / document title shown in row 1
    :param headers: Column header labels
    :param rows:    List of data rows (each row is a list matching headers order)
    :return:        BytesIO containing the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="2E6DA4")   # Deep blue
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font    = Font(name="Calibri", bold=True, size=14, color="1F3864")
    title_align   = Alignment(horizontal="center", vertical="center")

    cell_align    = Alignment(vertical="center", wrap_text=True)
    alt_fill      = PatternFill("solid", fgColor="EAF1FB")   # Light blue stripe

    thin_border   = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font   = title_font
    title_cell.fill   = PatternFill("solid", fgColor="D6E4F0")
    title_cell.alignment = title_align
    ws.row_dimensions[1].height = 28

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
    ws.row_dimensions[2].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=3):
        fill = alt_fill if row_idx % 2 == 1 else None
        ws.row_dimensions[row_idx].height = 18
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border    = thin_border
            if fill:
                cell.fill = fill

    # ── Auto-fit column widths ────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_values  = [str(header)] + [str(r[col_idx - 1]) for r in rows]
        max_length  = max((len(v) for v in col_values), default=10)
        adjusted    = min(max_length + 4, 50)          # cap at 50 chars
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # ── Freeze panes (header stays visible while scrolling) ───────────────────
    ws.freeze_panes = ws["A3"]

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
